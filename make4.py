import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook

class info:
    '''
    expected_result = [[101003487,13380,0], [101003473,1500,2]]
    1.[101003487,13380,0] : 1번째 생산정보, [101003473,1500,2] : 2번째 생산정보
    2.[제품코드, 요구물량, 작업전 품목교체시간]
    '''

    def __init__(self, load_ws, load_ws1, load_ws2):
        '''기준정보 읽어오기'''
        self.line_name = []
        self.pro_code = []
        self.pro_name = []
        self.ct = []
        self.multi_count = []
        self.row_ = []
        self.capa = []
        self.cluster = []
        self.girl = []

        for r in load_ws.rows:
            self.line_name.append(r[1].value)
            self.pro_code.append(r[2].value)
            self.pro_name.append(r[3].value)
            self.ct.append(r[4].value)
            self.multi_count.append(r[5].value)
            self.row_.append(r[6].value)
            self.capa.append(r[7].value)
            self.cluster.append(r[8].value)
            self.girl.append(r[9].value)

        self.girl = pd.DataFrame(self.girl[1:])
        BBB = pd.DataFrame(self.line_name[1:])
        AAA = pd.DataFrame(self.pro_code[1:])
        self.girl = pd.concat([BBB,pd.DataFrame(self.girl)], axis=1, ignore_index=True)
        self.girl = pd.concat([AAA, self.girl], axis=1, ignore_index=True)

        '''계획정보 읽어오기'''
        self.dm_line_name = []
        self.dm_pro_code = []
        self.dm_pro_demend = []
        self.dm_pro_seqence = []
        self.dm_pro_start = []

        for H in load_ws1.rows:
            self.dm_line_name.append(H[0].value)
            self.dm_pro_code.append(H[1].value)
            self.dm_pro_demend.append(H[2].value)
            self.dm_pro_seqence.append(H[3].value)
            self.dm_pro_start.append(H[4].value)

        self.dm_line_name = self.dm_line_name[1:]
        self.dm_pro_code = self.dm_pro_code[1:]
        self.dm_pro_demend = self.dm_pro_demend[1:]
        self.dm_pro_seqence = self.dm_pro_seqence[1:]
        self.dm_pro_start = self.dm_pro_start[1:]

        self.seq = {} #{'고속 22라인': 0, '고속 25라인': 1}
        for i in range(len(self.dm_line_name)):
            if self.dm_line_name[i] not in self.seq:
                self.seq[self.dm_line_name[i]] = self.dm_pro_start[i]


        #print(self.dm_line_name,'\n', self.dm_pro_code, '\n',
        #      self.dm_pro_demend, '\n', self.dm_pro_seqence)

        self.prev_list = []
        self.tray = []
        self.time = {}

        self.pro1 = []
        self.pro2 = []
        self.change_15 = []

        for U in load_ws2.rows:
            self.pro1.append(U[0].value)
            self.pro2.append(U[2].value)
            self.change_15.append(U[4].value)

        self.pro1 = self.pro1[1:]
        self.pro2 = self.pro2[1:]
        self.change_15 = self.change_15[1:]

    def sorting(self):
        '''
        [['고속 21라인' '101003486' '30500' '0']
         ['고속 21라인' '101003474' '2500' '1']
         ['고속 21라인' '101004606' '2500' '2']
         ['고속 21라인' '101000463' '11200' '3']
         ['고속 22라인' '101003487' '13380' '0']
         ['고속 22라인' '101003473' '1500' '1']
         ['고속 22라인' '101003388' '10500' '2']
         ['고속 22라인' '101003487' '7640' '3']
         ['고속 22라인' '101001977' '27350' '4']]
        '''
        run_line_list = []
        for v in self.dm_line_name:
            if v not in run_line_list:
                run_line_list.append(v)
        try:run_line_list.sort()
        except:pass

        self.tray = []
        for i in run_line_list:
            num = 0
            while True:
                #print('연산중...', np.array(self.tray))
                for j in range(len(self.dm_line_name)):
                    if self.dm_line_name[j] == i and self.dm_pro_seqence[j] == num:
                        if num != 0:  # 해당라인의 첫번째 생산이면...
                            '''Capa 및 교체시간 산출'''
                            capa, change_time = self.make_job_change_time(self.dm_line_name[j], self.dm_pro_code[j])

                            self.tray.append([self.dm_line_name[j],
                                              self.dm_pro_demend[j],
                                              self.dm_pro_seqence[j],
                                              capa, round(change_time,3),
                                              round(self.dm_pro_demend[j] / (capa / 10.5), 3,),
                                              self.dm_pro_code[j],
                                              self.pro_name[self.pro_code.index(self.dm_pro_code[j])]]
                                             )
                            self.prev_list = [self.dm_line_name[j], self.dm_pro_code[j]]
                        else:
                            self.prev_list = [self.dm_line_name[j], self.dm_pro_code[j]]
                            capa, _ = self.find_(self.dm_line_name[j], self.dm_pro_code[j])
                            self.tray.append([self.dm_line_name[j],
                                              self.dm_pro_demend[j],
                                              self.dm_pro_seqence[j],
                                              capa, 0,
                                              round(self.dm_pro_demend[j] / (capa / 10.5), 3),
                                              self.dm_pro_code[j],
                                              self.pro_name[self.pro_code.index(self.dm_pro_code[j])],]
                                             )

                        num += 1
                if num == self.dm_line_name.count(i) :
                    break

        time = self.cal_Retention_time(np.array(self.tray))
        #print(np.array(self.tray))

        return self.tray, time, self.seq, self.girl

    #생산요구량 수정
    def exchange_demend(self, line, id, new_demend):
        for i in range(len(self.dm_pro_code)):
            if self.dm_line_name[i] == line and self.dm_pro_seqence[i] == id:
                self.dm_pro_demend[i] = new_demend


    #생산순서 수정
    def exchange_seq(self, line, new_id):
        k=0
        while True:
            if self.dm_line_name[k] == line:
                for i in range(len(new_id)):
                    self.dm_pro_seqence[k + i] = int(new_id[i].strip())
                break
            else:
                k += 1

    # 팀내 전체 생산순서 수정
    def team_exchange_demend(self, An):
        '''
        self.dm_line_name = [] #라인명
        self.dm_pro_code = [] #제품코드
        self.dm_pro_seqence = [] #순서

        1.1팀 또는 2팀만 분류 하여 새로운 리스트 만들기 (고속 글자 시작 위치 파악하여 잘라내기)
        2.라인 중복 없애기
        3.라인별 개수 파악하여 리스트 만들기[3,4,4,4,5]
        4.시작위치 확인해가며 self.dm_pro_seqence 내 3.에서 만든 개수만큼 순서대로 랜덤 난수 발생시키고 저장
        '''

        temp_tray = [] # 라인 개수를 담을 그릇
        #print(self.dm_line_name[0])
        if An[0] == 1: #find not in 고속
            for h in self.dm_line_name:
                if '고속' not in h:
                    temp_tray.append(h)
        else :
            for h in self.dm_line_name:
                if '고속' in h:
                    temp_tray.append(h)

        #Make : ['굵은면 12라인 포장', '굵은면 13라인 포장', ...]
        bin = []
        for i in temp_tray:
            if i not in bin:
                bin.append(i)

        #bin 내 라인수 확인하여 몇개의 난수 발생시킬지 정하기 : [3,4,4,4,5]
        count_ = []
        for i in bin:
            count_.append(self.dm_line_name.count(i))

        #난수 리스트 만들기
        nan = []
        for j in range(len(bin)):
            nan_ = np.random.rand(count_[j])
            nan_ = [sorted(nan_).index(x) for x in nan_]
            nan.append(nan_)

        #난수 맵핑하기
        n = 0
        M = 0
        while True:
            if self.dm_line_name[n] == bin[M]:
                #nan = [0,1,2]
                for j in nan[M]:
                    self.dm_pro_seqence[n] = j
                    n += 1
                M += 1
            else :
                n += 1
            if n == len(self.dm_line_name) or M == len(bin):
                break


    #vlokup 구현
    def find_(self, line, code):
        for num in range(len(self.line_name)):
            if self.line_name[num] == line and self.pro_code[num] == code:
                return self.capa[num], self.cluster[num]


    #품목교체시간 산출
    def make_job_change_time(self, line_, code):
        change_time = None
        capa, clust_ = self.find_(line_, code)
        capa1, clust2_ = self.find_(self.prev_list[0], self.prev_list[1])

        if '고속' in line_:
            if clust_ == clust2_ :
                change_time = 1
            else: change_time = 2

        if '12' in line_:
            if clust_ == clust2_ :
                change_time = 1/2
            else: change_time = 1

        if '13' in line_:
            if clust_ == clust2_ :
                change_time = 1/3
            else: change_time = 1/2

        if '14' in line_:
            if clust_ == clust2_ :
                change_time = 0
            else: change_time = 1/3

        if '15' in line_:
            for i in range(len(self.pro1)):
                if self.prev_list[1] == self.pro1[i] and code == self.pro2[i]:
                    change_time = (self.change_15[i]/60)

        if '16' in line_:
            if clust_ == clust2_ :
                change_time = 1/3
            else: change_time = 1/2

        return capa, change_time


    def cal_Retention_time(self, tray):
        '''
        보유시간 산출하여 정상/연장 알맞게 그리기
        :return: [9, 10.5, 10.5, 10.5,  8,  8,  10.5, 10.5, 10.5, 6.7]
        '''
        long = np.array([9, 10.5, 10.5, 10.5, 8, 8, 10.5, 10.5, 10.5, 6.7])
        short = np.array([6.5, 8, 8, 8, 8, 8, 8, 8, 8, 8-(2.5)])
        half = np.array([9, 10.5, 10.5, 10.5, 8, 8, 8, 8, 8, 8-(10.5-6.7)])
        #print('long time : {}hr, half_time : {}hr, short_time : {}hr'
        #      .format(np.sum(long), np.sum(half), np.sum(short)))
        #print('-' * 70)

        run_line_list = []
        for v in self.dm_line_name:
            if v not in run_line_list:
                run_line_list.append(v)
        if len(run_line_list) > 1:
            run_line_list.sort()

        col_name = ['라인', '요구량', '순서', 'CAPA', '교체시간', '가동시간', 'code', '제품명']
        tray_pd = pd.DataFrame(tray, columns=col_name)

        for i in run_line_list:
            slice_list = tray_pd[tray_pd['라인'] == i].to_numpy()

            sum_ = sum(np.array(slice_list[:,4:6],dtype='float64'))
            sum_time = sum_[0] + sum_[1]

            if sum_time > np.sum(long):
                self.time[i] = '연장'
            elif sum_time < np.sum(long) :
                if sum_time > np.sum(half) :
                    self.time[i] = '반반'
                else :
                    self.time[i] = '정상'

        return self.time


class calcurate_line :
    def __init__(self, tray_, time_, sequence, girl_):
        self.tray = tray_
        self.time = time_
        self.sequence = sequence
        self.run_list = []
        self.change_list =[]
        self.have_time = []
        self.code = []
        self.total_sum = []
        self.girl = girl_


    def division(self, line):
        col_name = ['라인', '요구량', '순서', 'CAPA', '교체시간', '가동시간', 'code', '제품명']
        tray_pd = pd.DataFrame(self.tray, columns=col_name)
        slice = tray_pd[tray_pd['라인'] == line].to_numpy()

        return slice


    def make_time_list(self, list):
        run_line_list = []
        capa_list = []
        for _, amount, seq, capa, change, run_time, code, prod in list:
            if code not in run_line_list:
                run_line_list.append(code)
                capa_list.append(capa/10.5)

        self.change_tray = []
        for _ in range(len(run_line_list)):
            self.change_tray.append([0.0] * 10)

        self.capa_tray = []

        for i in range(len(run_line_list)):
            self.capa_tray.append([capa_list[i]] * 10)

        self.run_list = pd.DataFrame(self.change_tray,  index=run_line_list)
        self.change_list = pd.DataFrame(self.change_tray,  index=run_line_list)
        self.code = run_line_list


    def have(self, list):
        '''
        해당조 self.sequence 내 잔여 여유시간을 산출하기
        '''

        self.have_time = []
        if '고속' in list[0][0] :
            # 보유시간(연장/정상) 가져오기
            if self.time[list[0][0]] == '연장':
                self.have_time = [9.0, 10.5, 10.5, 10.5, 8, 8, 10.5, 10.5, 10.5, 6.7]
            elif self.time[list[0][0]] == '정상':
                self.have_time = [6.5, 8, 8, 8, 8, 8, 8, 8, 8, 8 - (10.5 - 6.7)]
            elif self.time[list[0][0]] == '반반':
                self.have_time = [9.0, 10.5, 10.5, 10.5, 8, 8, 8, 8, 8, 8 - (10.5 - 6.7)]
        else:
            # 보유시간(연장/정상) 가져오기
            if self.time[list[0][0]] == '연장':
                self.have_time = [9.5, 10.5, 10.5, 10.5, 8, 8, 10.5, 10.5, 10.5, 8]
            elif self.time[list[0][0]] == '정상':
                self.have_time = [7, 8, 8, 8, 8, 8, 8, 8, 8, 8 - (2.5)]
            elif self.time[list[0][0]] == '반반':
                self.have_time = [9.5, 10.5, 10.5, 10.5, 8, 8, 8, 8, 8, 8 - (2.5)]


        try:
            return self.have_time[self.sequence] - (self.run_list[self.sequence].sum()+self.change_list[self.sequence].sum())
        except:
            self.sequence = 9


    def change_T(self, list, code, change, before_code):
        '''
        교체시간을 self.change_list 안에 넣기
        '''
        if self.sequence != 10:
            total_T = self.have(list) #total_T : 현재 sequence에 여유시간
            '''
            폼목교체 조건 문
            '''
            if total_T >= change:  # 남은 시간이 넣어야하는 교체시간 보다 클경우
                if self.have_time[self.sequence] == 8 and \
                        self.sequence in [1,3,5,7,9] and \
                        total_T == 8:
                    #print('hey.......', self.sequence, total_T, self.have_time[self.sequence])
                    self.change_list.loc[code][self.sequence] = 0
                else:
                    # 교체시간 넣기
                    self.change_list.loc[code][self.sequence] = change #[어떤제품인지][몇번째조인지] = 교체시간 삽입
            else:
                if total_T - change < 0 :
                    self.run_T(list, before_code, total_T, change)
                    total_T = self.have(list)  # total_T : 현재 sequence에 여유시간
                    if self.have_time[self.sequence] == 8 and \
                            self.sequence in [1, 3, 5, 7, 9] and \
                            total_T == 8:
                        #print('hey.......', self.sequence, total_T, self.have_time[self.sequence])
                        self.change_list.loc[code][self.sequence] = 0
                    else :
                        self.change_list.loc[code][self.sequence] = change
                else:
                    self.change_list.loc[code][self.sequence] = change  # [어떤제품인지][몇번째조인지] = 교체시간 삽입
                #self.run_T(list, before_code, total_T)

            '''
            교체시간을 넣은 후에 해당되는 조 self.sequence 내 잔여 여유시간이 0.5 이하일때
            전에 제품에 잔여시간을 가동시간으로 넣기
            '''
            #total_T = self.have(list) #total_T : 현재 sequence에 여유시간
            #if total_T < change: # 남은 시간이 넣어야하는 교체시간 보다 작을 경우


    def run_T(self, list, code, run_time, change):
        while self.sequence < 10:
            # 해당 조에 남은 시간 확인
            try:
                total_T = self.have(list) #total_T : 현재 sequence에 여유시간
            except :
                break

            if total_T > run_time:
                self.run_list.loc[code][self.sequence] += float(run_time)
                run_time = 0
                if total_T - run_time < 0.2:
                    self.sequence += 1

            elif total_T == run_time:
                self.run_list.loc[code][self.sequence] += float(run_time)
                run_time = 0
                self.sequence += 1

            elif total_T < run_time:
                self.run_list.loc[code][self.sequence] += float(total_T)
                run_time = run_time - total_T
                if run_time < change :
                    run_time = 0
                self.sequence += 1

            if 0 >= run_time:
                break


    def put_data_in_line(self, list):
        before_code = 100000000
        for _, amount, seq, capa, change, run_time, code, prod in list:
            '''
            _:라인, amount:요구량, seq:순서, capa:capa, change:교체시간, run_time:가동시간, prod:제품명
            '''
            if self.sequence != 10:
                # 교체시간 넣기
                self.change_T(list, code, change, before_code)
                 # 생산시간 넣기
                self.run_T(list, code, run_time, change)
                before_code = code

        #print('\n','-' * 70)
        #print(list) #나중에 켜세요
        #print(self.run_list, '\n', self.change_list)

    def name(self, Result):
        name = []
        for i in np.array(Result.index):
            j = 0
            while True:
                if self.tray[j][6] == i:
                    name.append(self.tray[j][7])
                    break
                else:
                    j+=1

        return name


    def run(self, seq_):

        for i, line in enumerate(self.time):  # 0, 고속 21라인형태 임
            self.sequence = seq_[line]

            if self.sequence != 10:
                list = self.division(line)
                self.make_time_list(list)
                self.put_data_in_line(list)

                A = np.array(self.capa_tray)
                B = np.array(self.run_list)
                sum_ = np.round(np.array(A * B, dtype='int'), -1)


               # print('-' * 70)
                #print(pd.DataFrame(sum_, index=self.code)) #나중에 켜세요
                #print('*' * 70, '\n') #나중에 켜세요

                df = pd.DataFrame(sum_, index=self.code)
                new_ = []
                for j in range(len(sum_)):
                    new_.append(line)
                df = pd.concat([pd.DataFrame(new_, index=self.code), df], axis=1, ignore_index=True)

                if i == 0:
                    self.total_sum = df
                else:
                    self.total_sum = pd.concat([self.total_sum, df], ignore_index=False)

        return self.total_sum


    def time_valid(self):
        '''
        self.total_sum을 보고 팀내 조별 필요 인원수가 TO인원을 넘어설 시
        생산순서를 랜덤으로 바꾸기
        :return: 문제가 없을 시 1 return, 문제가 있을 시 0 return
        '''

        temp = self.total_sum.reset_index().values
        slow = [] #'1팀 생산현황 넣기'
        fast = [] #'2팀 생산현황 넣기'

        for i in temp:
            if '고속' in i[1] :
                fast.append(i)
            else :
                slow.append(i)
        list_problem = []
        hm = 0

        for h in [slow, fast]:
            slow = np.array(h).tolist()
            slow_df = pd.DataFrame(slow)
            line = slow_df.pop(1)

            for j in line.drop_duplicates().tolist():  # 인덱스를 중복제거 후 리스트로 변환
                AAA = pd.DataFrame(slow)
                BBB = AAA[1] == j
                impor = AAA[BBB].sum().values.tolist()[-10:]

                for num in range(len(slow)):

                    if slow[num][1] == j:  # j 라인명과 slow 리스트 두번째 columns 가 같으면
                        # print('{}라인에 제품 {}를 기준으로 투입인원 넣기'.format(j, slow[num][0]))
                        ls_LINE = self.girl[1] == j
                        is_code = self.girl[0] == slow[num][0]

                        x = self.girl[ls_LINE&is_code][2].values[0]  # 얻어진 인원 수

                        for jo in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
                            '''
                            나중에 jo(조)에 생산이 없을 시 To 인원 빼기 보강필요
                            라인별 검사하여 빼기 기능 넣기 (예약)
                            '''
                            if impor[jo-2] == 0: #생산이 없으면
                                slow[num][jo] = -5 #인원수를 뺀다
                            elif slow[num][jo] != 0:
                                slow[num][jo] = x

            A = pd.DataFrame(slow)

            del A[1]
            del A[0]
            A = A.values.tolist()

            slow = pd.DataFrame(A, index = line.values.tolist())

            tray = []
            for v in line.drop_duplicates().tolist():
                if len(slow.loc[v]) == 10:
                    tray.append(slow.loc[v].values.tolist())
                else:
                    des = slow.loc[v].describe()
                    _tray = des.loc['min'].values.tolist()
                    cont = des.iloc[7].values.tolist()
                    temp2 = []
                    for k in range(len(_tray)): #쫄병 처럼 인원수가 - 값을 가지면 min 값 반환
                        if _tray[k] < 0 :
                            temp2.append(_tray[k])
                        else:
                            temp2.append(cont[k])

                    tray.append(temp2)


            B = pd.DataFrame(tray).sum().values.tolist()

            if hm == 0:
                max_allow = 4 #1팀일때
            else :
                max_allow = 2 #2팀일때

            if max(B) <= max_allow:
                list_problem.append(0)
            else:
                list_problem.append(1)
            hm += 1

        return list_problem


    def export_valid(self, Result):
        #생산순서에 따른 제품명 리스트 가져오기
        name_list = self.name(Result)

        #Result 데이터 프레임 맨 오른쪽에 제품명 넣기
        Result[11] = name_list

        #제품명 열 이름 name으로 변경
        Result.rename(columns={11: 'name'}, inplace=True)

        #21호와 22호에 수출 제품만 sorting
        _21_sort = Result[Result[0].isin(['고속 21라인 포장'])]
        _21_sort = _21_sort[_21_sort['name'].str.contains('수출')].values.tolist()[0][1:-1]

        _22_sort = Result[Result[0].isin(['고속 22라인 포장'])]
        _22_sort = _22_sort[_22_sort['name'].str.contains('수출')].values.tolist()[0][1:-1]

        #21호와 22호에 수출이 있을 땐 같이 있어야함. 그렇지 못할 경우 x에 1을 반환
        x = 0
        for i in range(10):
            if _21_sort[i] > 0 and _22_sort[i] <= 0:
                x = 1
        for i in range(10):
            if _22_sort[i] > 0 and _21_sort[i] <= 0:
                x = 1

        print(_21_sort)
        print(_22_sort)
        print(x)
        return x


if __name__ == '__main__':

    load_wb = load_workbook("info.xlsx", data_only=True)
    load_wb1 = load_workbook("plan.xlsx", data_only=True)
    load_wb2 = load_workbook("one_five_line_change.xlsx", data_only=True)

    load_ws = load_wb['Sheet1']
    load_ws1 = load_wb1['Sheet1']
    load_ws2 = load_wb2['Sheet1']

    F = info(load_ws, load_ws1, load_ws2)
    print('기준정보 런칭 완료')
    '''time_ : {'고속 21라인': '정상', '고속 22라인': '반반'}'''
    tray_, time_, seq_, girl = F.sorting()
    print('기준정보 정리 완료')

    start, num = 0, 0
    while True:

        if start != 0:
            Answer = input('\n수정하시겠습니까?? 1:생산량 조정, 2:생산순서 조정, 3:종료 = ')
            os.system('cls')
            if Answer == '1':
                answer = input('세부정보를 입력하세요(따음표로 구분).. 예) 고속 22라인, 0(제품), 50000(요구량)\n')
                answer_ = answer.split(',')
                # 생산요구량 수정
                F.exchange_demend(line=answer_[0], id=int(answer_[1]), new_demend=int(answer_[2]))
                tray_, time_, seq_, girl = F.sorting()

            if Answer == '2':
                answer = input('세부정보를 입력하세요(따음표로 구분).. 예) 고속 22라인, 0, 2, 1, 3\n')
                answer_ = answer.split(',')
                bin = answer_[1:]
                print(bin,'*******=====\n')
                # 생산순서 수정
                F.exchange_seq(line=answer_[0], new_id=bin)
                tray_, time_, seq_, girl = F.sorting()

            if Answer == '3':
                #생산계획서에 0 은  ' ' 로 변환
                Result = Result.replace({0:''})

                #새로 만들어진 제품 순서별 제품 명을 반환
                name_list = C.name(Result)

                #제품 이름을 12번째 열에 넣기
                Result[11] = name_list

                #주간조 표시하기 위해 븬 리스트 만들기
                empty = ['']*len(name_list)

                #븬 리스트를 맨 뒤열에 넣기
                Result[12] = empty
                Result[13] = empty
                Result[14] = empty
                Result[15] = empty
                Result[16] = empty

                #열 순서 재배치
                Result = Result[[11,0,1,12,2,3,13,4,5,14,6,7,15,8,9,16,10]]

                Result.reset_index()
                Result.to_excel('plan2.xlsx', sheet_name='new_plan')
                break

        C = calcurate_line(tray_, time_, 0, girl)

        # 품목교체 등을 고려하여 요구량을 계획서에 뿌려줌
        Result = C.run(seq_)

        # 계획서에 제약조건 확인
        An = C.time_valid()

        # 2팀 수출제품 문제가 있으면 1, 없으면 0을 반환
        export_error = 0
        export_error = C.export_valid(Result)
        if export_error == 1:
            An[1] = 1

        '''
        제품별 전체 합이 0보다 같거나 작으면 1을 반환하는 함수 만들어야함
        *.제품이 미 할당 되거나 (-) 값을 가지는것을 방지
        '''

        # [0,1] 1팀, 2팀 위치로 1이 있으면 문제가 있다고 판단하여 다시 순서 섞기
        if 1 in An:
            start = 0
            if 1 == An[0]:
                print('1팀 순서 흔들기....')
                F.team_exchange_demend(An)
                tray_, time_, seq_, girl = F.sorting()
            else :
                print('2팀 순서 흔들기....')
                F.team_exchange_demend(An)
                tray_, time_, seq_, girl = F.sorting()
        else:
            start += 1

        print('{}번째 연산..'.format(num))
        print('-' * 70)
        num += 1

    print(Result)